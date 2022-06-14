from openpyxl import load_workbook
from base import analysis_mcasem1, analysis_mcasem2, analysis_mcasem3


def filterClass(name, column, wb, ws1, ws2):
    rows = [1]
    for i in range(2, ws1.max_row+1):
        if (ws1.cell(row=i, column=column).value).lower().find(name) != -1:
            rows.append(i)

    j = 1
    for row in rows:
        for i in range(1, ws1.max_column+1):
            ws2.cell(row=j, column=i).value = ws1.cell(
                row=row, column=i).value
        j += 1
    return wb


def class_analysis(excel, class_name):
    print(excel)
    if class_name == "mca_sem1":
        wb = load_workbook(excel)
        ws1 = wb['Sheet1']

        wb.create_sheet('morning')
        wb.create_sheet('afternoon')

        ws2 = wb['morning']
        wb = filterClass("morning", 85, wb, ws1, ws2)

        ws3 = wb['afternoon']
        wb = filterClass("afternoon", 85, wb, ws1, ws3)

        wb.save(excel)

        analysis_mcasem1.genrateAnalysis_MCAsem1(
            excel, "morning", "SEM1 ANALYSIS Reg. Morning")

        analysis_mcasem1.genrateAnalysis_MCAsem1(
            excel, "afternoon", "SEM1 ANALYSIS Reg. Afternoon")

        return excel

    elif class_name == "mca_sem2":
        wb = load_workbook(excel)
        ws1 = wb['Sheet1']

        wb.create_sheet('morning')
        wb.create_sheet('afternoon')

        ws2 = wb['morning']
        wb = filterClass("morning", 105, wb, ws1, ws2)

        ws3 = wb['afternoon']
        wb = filterClass("afternoon", 105, wb, ws1, ws3)

        wb.save(excel)

        analysis_mcasem2.genrateAnalysis_MCAsem2(
            excel, "morning", "SEM2 ANALYSIS Reg. Morning")

        analysis_mcasem2.genrateAnalysis_MCAsem2(
            excel, "afternoon", "SEM2 ANALYSIS Reg. Afternoon")

        return excel

    elif class_name == "mca_sem3":
        wb = load_workbook(excel)
        ws1 = wb['Sheet1']

        wb.create_sheet('morning')
        wb.create_sheet('afternoon')

        ws2 = wb['morning']
        wb = filterClass("morning", 86, wb, ws1, ws2)

        ws3 = wb['afternoon']
        wb = filterClass("afternoon", 86, wb, ws1, ws3)

        wb.save(excel)

        analysis_mcasem3.genrateAnalysis_MCAsem3(
            excel, "morning", "SEM3 ANALYSIS Reg. Morning")

        analysis_mcasem3.genrateAnalysis_MCAsem3(
            excel, "afternoon", "SEM3 ANALYSIS Reg. Afternoon")

        return excel
