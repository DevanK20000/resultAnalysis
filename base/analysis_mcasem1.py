from django.conf import settings
from numpy import double
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.chart import BarChart, Reference, Series
import pandas as pd


def create_Boilerplate(wb, class_name):
    ws = wb[class_name]
    ws.column_dimensions['A'].width = 40

    ws.merge_cells('A2:H2')
    ws['A2'] = 'Vivekanand Education Society Institute of Technology'

    ws.merge_cells('A3:H3')
    ws['A3'] = 'Master of Computer Applications'

    ws.merge_cells('A4:H4')
    ws['A4'] = 'Result Analysis –Ist year-Sem 1'

    ws.merge_cells('A7:H7')
    ws['A7'] = 'OVERALL SUMMARY OF THE SEMESTER'

    ws['A2'].font = Font(size=12, bold=True)
    ws['A3'].font = Font(size=12, bold=True)
    ws['A4'].font = Font(size=12, bold=True)
    ws['A7'].font = Font(size=14, bold=True)

    ws['A2'].alignment = Alignment(horizontal='center')
    ws['A3'].alignment = Alignment(horizontal='center')
    ws['A4'].alignment = Alignment(horizontal='center')
    ws['A7'].alignment = Alignment(horizontal='center')

    return wb


def getAverage(column, ws):
    average = 0
    for i in range(2, ws.max_row+1):
        average += int(ws.cell(row=i, column=column).value)
    average = round(average/(ws.max_row-1))
    return average


def getPassed(column, ws):
    count = 0
    for i in range(2, ws.max_row+1):
        if str(ws.cell(row=i, column=column).value) != "--" and str(ws.cell(row=i, column=column).value) != "A" and str(ws.cell(row=i, column=column).value).find("F") == -1:
            count += 1
    return count


def getTotalPFA(column, g, ws):
    total = 0
    for i in range(2, ws.max_row+1):
        if ws.cell(row=i, column=column).value == g:
            total += 1
    return total


def checkAppeared(column, ws):
    count = 0
    for i in range(2, ws.max_row+1):
        if str(ws.cell(row=i, column=column).value).find("A") == -1:
            count += 1
    return count


def getGradeCount(column, grade, ws):
    count = 0
    for i in range(2, ws.max_row+1):
        if ws.cell(row=i, column=column).value == grade:
            count += 1
    return count


def percentageOfStudentsWith60amdAbovePercentage(column, ws):
    count = 0
    for i in range(2, ws.max_row+1):
        if (ws.cell(row=i, column=column).value) != "--":
            if double(ws.cell(row=i, column=column).value)*7.4+12 >= 60:
                count += 1
    return (count/(ws.max_row-1))*100


def percentageOfStudentsBelow60(column, ws):
    count = 0
    for i in range(2, ws.max_row+1):
        if (ws.cell(row=i, column=column).value) != "--":
            if double(ws.cell(row=i, column=column).value)*7.4+12 < 60:
                count += 1
    return count/(ws.max_row-1)*100


def overall_summary_of_the_semester(wb, source, class_name):
    ws = wb[source]
    ws2 = wb[class_name]

    ws2.append(['Subjects', 'MFCS', 'MFCS-TW', 'AJAVA', 'AJL',
               'ADBMS', 'ADBMSL', 'SPM', 'SPM', 'DSL', 'WTL', 'MP', 'Total'])
    ws2.append(['Marks', '100', '25', '100', '75', '100',
               '75', '100', '25', '100', '100', '50', '850'])
    ws2.append(['Total Students', ws.max_row-1, ws.max_row-1, ws.max_row-1, ws.max_row-1, ws.max_row -
               1, ws.max_row-1, ws.max_row-1, ws.max_row-1, ws.max_row-1, ws.max_row-1, ws.max_row-1])
    ws2.append(['Total Apeared', checkAppeared(6, ws), checkAppeared(12, ws), checkAppeared(19, ws), checkAppeared(26, ws), checkAppeared(
        33, ws), checkAppeared(40, ws), checkAppeared(47, ws), checkAppeared(53, ws), checkAppeared(60, ws), checkAppeared(67, ws), checkAppeared(73, ws)])
    ws2.append(['Average Marks', getAverage(7, ws), getAverage(13, ws), getAverage(20, ws), getAverage(27, ws), getAverage(
        34, ws), getAverage(41, ws), getAverage(48, ws), getAverage(54, ws), getAverage(61, ws), getAverage(68, ws), getAverage(74, ws)])
    ws2.append(["Total Passed", getPassed(7, ws), getPassed(13, ws), getPassed(20, ws), getPassed(27, ws), getPassed(
        34, ws), getPassed(41, ws), getPassed(48, ws), getPassed(54, ws), getPassed(61, ws), getPassed(68, ws), getPassed(74, ws)])
    ws2['M13'] = getTotalPFA(83, "P", ws)
    ws2.append(["Total Failed", (ws2["B11"].value-ws2["B13"].value), (ws2["C11"].value-ws2["C13"].value), (ws2["D11"].value-ws2["D13"].value), (ws2["E11"].value-ws2["E13"].value), (ws2["F11"].value-ws2["F13"].value),
               (ws2["G11"].value-ws2["G13"].value), (ws2["H11"].value-ws2["H13"].value), (ws2["I11"].value-ws2["I13"].value), (ws2["J11"].value-ws2["J13"].value), (ws2["K11"].value-ws2["K13"].value), (ws2["L11"].value-ws2["L13"].value)])
    ws2['M14'] = (ws.max_row-1)-ws2['M13'].value
    ws2.append(["Total Absent", (ws2["B10"].value-ws2["B11"].value), (ws2["C10"].value-ws2["C11"].value), (ws2["D10"].value-ws2["D11"].value), (ws2["E10"].value-ws2["E11"].value), (ws2["F10"].value-ws2["F11"].value),
               (ws2["G10"].value-ws2["G11"].value), (ws2["H10"].value-ws2["H11"].value), (ws2["I10"].value-ws2["I11"].value), (ws2["J10"].value-ws2["J11"].value), (ws2["K10"].value-ws2["K11"].value), (ws2["L10"].value-ws2["L11"].value)])
    ws2['M15'] = getTotalPFA(83, "A", ws)
    ws2.append(["Percentage Passed", ws2['B13'].value/ws2["B10"].value*100, ws2['C13'].value/ws2["C10"].value*100, ws2['C13'].value/ws2["C10"].value*100, ws2['D13'].value/ws2["D10"].value*100, ws2['E13'].value/ws2["E10"].value*100, ws2['F13'].value /
               ws2["F10"].value*100, ws2['G13'].value/ws2["G10"].value*100, ws2['H13'].value/ws2["H10"].value*100, ws2['I13'].value/ws2["I10"].value*100, ws2['J13'].value/ws2["J10"].value*100, ws2['K13'].value/ws2["K10"].value*100, ws2['L13'].value/ws2["L10"].value*100])
    ws2['M16'] = ws2['M13'].value/(ws.max_row-1)*100
    ws2.append(["Percentage Failed", ws2['B14'].value/ws2["B10"].value*100, ws2['C14'].value/ws2["C10"].value*100, ws2['C14'].value/ws2["C10"].value*100, ws2['D14'].value/ws2["D10"].value*100, ws2['E14'].value/ws2["E10"].value*100, ws2['F14'].value /
               ws2["F10"].value*100, ws2['G14'].value/ws2["G10"].value*100, ws2['H14'].value/ws2["H10"].value*100, ws2['I14'].value/ws2["I10"].value*100, ws2['J14'].value/ws2["J10"].value*100, ws2['K14'].value/ws2["K10"].value*100, ws2['L14'].value/ws2["L10"].value*100])
    ws2['M17'] = ws2['M14'].value/(ws.max_row-1)*100
    ws2.append(["Number of students with >= 80% ie =Grade O", getGradeCount(9, 'O', ws), getGradeCount(15, 'O', ws), getGradeCount(22, 'O', ws), getGradeCount(29, 'O', ws), getGradeCount(
        36, 'O', ws), getGradeCount(43, 'O', ws), getGradeCount(50, 'O', ws), getGradeCount(56, 'O', ws), getGradeCount(63, 'O', ws), getGradeCount(70, 'O', ws), getGradeCount(76, 'O', ws)])
    ws2.append(["Number of students with >= 75% ie =Grade A", getGradeCount(9, 'A', ws), getGradeCount(15, 'A', ws), getGradeCount(22, 'A', ws), getGradeCount(29, 'A', ws), getGradeCount(
        36, 'A', ws), getGradeCount(43, 'A', ws), getGradeCount(50, 'A', ws), getGradeCount(56, 'A', ws), getGradeCount(63, 'A', ws), getGradeCount(70, 'A', ws), getGradeCount(76, 'A', ws)])
    ws2.append(["Number of students with >= 70% ie =Grade B", getGradeCount(9, 'B', ws), getGradeCount(15, 'B', ws), getGradeCount(22, 'B', ws), getGradeCount(29, 'B', ws), getGradeCount(
        36, 'B', ws), getGradeCount(43, 'B', ws), getGradeCount(50, 'B', ws), getGradeCount(56, 'B', ws), getGradeCount(63, 'B', ws), getGradeCount(70, 'B', ws), getGradeCount(76, 'B', ws)])
    ws2.append(["Number of students with >= 60% ie =Grade C", getGradeCount(9, 'C', ws), getGradeCount(15, 'C', ws), getGradeCount(22, 'C', ws), getGradeCount(29, 'C', ws), getGradeCount(
        36, 'C', ws), getGradeCount(43, 'C', ws), getGradeCount(50, 'C', ws), getGradeCount(56, 'C', ws), getGradeCount(63, 'C', ws), getGradeCount(70, 'C', ws), getGradeCount(76, 'C', ws)])
    ws2.append(["Number of students with >= 55% ie =Grade D", getGradeCount(9, 'D', ws), getGradeCount(15, 'D', ws), getGradeCount(22, 'D', ws), getGradeCount(29, 'D', ws), getGradeCount(
        36, 'D', ws), getGradeCount(43, 'D', ws), getGradeCount(50, 'D', ws), getGradeCount(56, 'D', ws), getGradeCount(63, 'D', ws), getGradeCount(70, 'D', ws), getGradeCount(76, 'D', ws)])
    ws2.append(["Number of students with >= 50% ie =Grade E", getGradeCount(9, 'E', ws), getGradeCount(15, 'E', ws), getGradeCount(22, 'E', ws), getGradeCount(29, 'E', ws), getGradeCount(
        36, 'E', ws), getGradeCount(43, 'E', ws), getGradeCount(50, 'E', ws), getGradeCount(56, 'E', ws), getGradeCount(63, 'E', ws), getGradeCount(70, 'E', ws), getGradeCount(76, 'E', ws)])
    ws2.append(["Number of students with >= 45% ie =Grade P", getGradeCount(9, 'P', ws), getGradeCount(15, 'P', ws), getGradeCount(22, 'P', ws), getGradeCount(29, 'P', ws), getGradeCount(
        36, 'P', ws), getGradeCount(43, 'P', ws), getGradeCount(50, 'P', ws), getGradeCount(56, 'P', ws), getGradeCount(63, 'P', ws), getGradeCount(70, 'P', ws), getGradeCount(76, 'P', ws)])
    ws2.append(["Number of students with <= 45% ie =Grade F", getGradeCount(9, 'F', ws), getGradeCount(15, 'F', ws), getGradeCount(22, 'F', ws), getGradeCount(29, 'F', ws), getGradeCount(
        36, 'F', ws), getGradeCount(43, 'F', ws), getGradeCount(50, 'F', ws), getGradeCount(56, 'F', ws), getGradeCount(63, 'F', ws), getGradeCount(70, 'F', ws), getGradeCount(76, 'F', ws)])
    ws2.append(["Total Number of students with 60% and above",
                ws2['B18'].value+ws2['B19'].value +
                ws2['B20'].value+ws2['B21'].value,
                ws2['C18'].value+ws2['C19'].value +
                ws2['C20'].value+ws2['C21'].value,
                ws2['D18'].value+ws2['D19'].value +
                ws2['D20'].value+ws2['D21'].value,
                ws2['E18'].value+ws2['E19'].value +
                ws2['E20'].value+ws2['E21'].value,
                ws2['F18'].value+ws2['F19'].value +
                ws2['F20'].value+ws2['F21'].value,
                ws2['G18'].value+ws2['G19'].value +
                ws2['G20'].value+ws2['G21'].value,
                ws2['H18'].value+ws2['H19'].value +
                ws2['H20'].value+ws2['H21'].value,
                ws2['I18'].value+ws2['I19'].value +
                ws2['I20'].value+ws2['I21'].value,
                ws2['J18'].value+ws2['J19'].value +
                ws2['J20'].value+ws2['J21'].value,
                ws2['K18'].value+ws2['K19'].value +
                ws2['K20'].value+ws2['K21'].value,
                ws2['L18'].value+ws2['L19'].value+ws2['L20'].value+ws2['L21'].value])

    ws2.append(["% of students with 60% and above",
                (ws2["B26"].value/ws2["B10"].value *
                 100)if ws2["B10"].value != 0 else 0,
                (ws2["C26"].value/ws2["C10"].value *
                 100)if ws2["C10"].value != 0 else 0,
                (ws2["D26"].value/ws2["D10"].value *
                 100)if ws2["D10"].value != 0 else 0,
                (ws2["E26"].value/ws2["E10"].value *
                 100)if ws2["E10"].value != 0 else 0,
                (ws2["F26"].value/ws2["F10"].value *
                 100)if ws2["F10"].value != 0 else 0,
                (ws2["G26"].value/ws2["G10"].value *
                 100)if ws2["G10"].value != 0 else 0,
                (ws2["H26"].value/ws2["H10"].value *
                 100)if ws2["H10"].value != 0 else 0,
                (ws2["I26"].value/ws2["I10"].value *
                 100)if ws2["I10"].value != 0 else 0,
                (ws2["J26"].value/ws2["J10"].value *
                 100)if ws2["J10"].value != 0 else 0,
                (ws2["K26"].value/ws2["K10"].value *
                 100)if ws2["K10"].value != 0 else 0,
                (ws2["L26"].value/ws2["L10"].value*100)if ws2["L10"].value != 0 else 0])

    ws2.append(["Total Number of students below 60%",
                ws2["B10"].value-ws2['B26'].value,
                ws2["C10"].value-ws2['C26'].value,
                ws2["D10"].value-ws2['D26'].value,
                ws2["E10"].value-ws2['E26'].value,
                ws2["F10"].value-ws2['F26'].value,
                ws2["G10"].value-ws2['G26'].value,
                ws2["H10"].value-ws2['H26'].value,
                ws2["I10"].value-ws2['I26'].value,
                ws2["J10"].value-ws2['J26'].value,
                ws2["K10"].value-ws2['K26'].value,
                ws2["L10"].value-ws2['L26'].value])
    ws2.append(["% of students below 60%",
                (ws2["B28"].value/ws2['B26'].value *
                 100)if ws2["B26"].value != 0 else 0,
                (ws2["C28"].value/ws2['C26'].value *
                 100)if ws2["C26"].value != 0 else 0,
                (ws2["D28"].value/ws2['D26'].value *
                 100)if ws2["D26"].value != 0 else 0,
                (ws2["E28"].value/ws2['E26'].value *
                 100)if ws2["E26"].value != 0 else 0,
                (ws2["F28"].value/ws2['F26'].value *
                 100)if ws2["F26"].value != 0 else 0,
                (ws2["G28"].value/ws2['G26'].value *
                 100)if ws2["G26"].value != 0 else 0,
                (ws2["H28"].value/ws2['H26'].value *
                 100)if ws2["H26"].value != 0 else 0,
                (ws2["I28"].value/ws2['I26'].value *
                 100)if ws2["I26"].value != 0 else 0,
                (ws2["J28"].value/ws2['J26'].value *
                 100)if ws2["J26"].value != 0 else 0,
                (ws2["K28"].value/ws2['K26'].value *
                 100)if ws2["K26"].value != 0 else 0,
                (ws2["L28"].value/ws2['L26'].value*100)if ws2["L26"].value != 0 else 0])

    ws2.append(["Grand Total",
                ws2['B26'].value+ws2['B28'].value,
                ws2['C26'].value+ws2['C28'].value,
                ws2['D26'].value+ws2['D28'].value,
                ws2['E26'].value+ws2['E28'].value,
                ws2['F26'].value+ws2['F28'].value,
                ws2['G26'].value+ws2['G28'].value,
                ws2['H26'].value+ws2['H28'].value,
                ws2['I26'].value+ws2['I28'].value,
                ws2['J26'].value+ws2['J28'].value,
                ws2['K26'].value+ws2['K28'].value,
                ws2['L26'].value+ws2['L28'].value])

    ws2.append([""])
    ws2.append(["Total Student appeared"])
    ws2.append(["passsed", getTotalPFA(83, "P", ws)])
    ws2.append(["failed", getTotalPFA(83, "F", ws)])
    ws2.append(["Absent", getTotalPFA(83, "A", ws)])
    ws2.append(["Total passing percentage",
               ws2['M13'].value/(ws.max_row-1)*100])
    ws2.append(["Total failing percentage",
               ws2['M14'].value/(ws.max_row-1)*100])
    ws2.append(["% of students with 60% and above",
               percentageOfStudentsWith60amdAbovePercentage(82, ws)])
    ws2.append(["% of students with below 60%", percentageOfStudentsBelow60(
        82, ws), "", "Total Absent percentage", "", ws2['B35'].value/(ws.max_row-1)*100])
    ws2.append([""])
    ws2.append(['Subjects', 'MFCS', 'MFCS-TW', 'AJAVA', 'AJL',
               'ADBMS', 'ADBMSL', 'SPM', 'SPM', 'DSL', 'WTL', 'MP', 'Total'])
    ws2.append(["% of students with 60% and above", ws2['B27'].value,
                ws2['C27'].value,
                ws2['D27'].value,
                ws2['E27'].value,
                ws2['F27'].value,
                ws2['G27'].value,
                ws2['H27'].value,
                ws2['I27'].value,
                ws2['J27'].value,
                ws2['K27'].value,
                ws2['L27'].value])
    ws2.append(["% of students below 60%", ws2['B28'].value,
                ws2['C29'].value,
                ws2['D29'].value,
                ws2['E29'].value,
                ws2['F29'].value,
                ws2['G29'].value,
                ws2['H29'].value,
                ws2['I29'].value,
                ws2['J29'].value,
                ws2['K29'].value,
                ws2['L29'].value])
    return wb


def findTop10Rankers(ws):
    top10 = {
        "seat": [],
        "name": [],
        "total": [],
        "outof": [],
        "GPA": [],
    }
    for i in range(2, ws.max_row):
        top10["seat"].append(ws.cell(row=i, column=2).value)
        top10["name"].append(ws.cell(row=i, column=3).value)
        top10["total"].append(ws.cell(row=i, column=79).value.split("/")[0])
        top10["outof"].append(ws.cell(row=i, column=79).value.split("/")[1])
        top10["GPA"].append(ws.cell(row=i, column=82).value)

    top10 = pd.DataFrame(top10)
    top10 = top10.sort_values(
        by=["total", "GPA"], ignore_index=True, ascending=False)
    return top10


def subjectRankers(column, subjects, ws):
    subjectRankers = {
        "seat": [],
        "name": [],
        "subject": [],
    }

    for col in column:
        temp = {
            "seat": [],
            "name": [],
            "marks": [],
            "GPA": [],
        }
        for i in range(2, ws.max_row):
            temp["seat"].append(ws.cell(row=i, column=2).value)
            temp["name"].append(ws.cell(row=i, column=3).value)
            try:
                temp["marks"].append(int(ws.cell(row=i, column=col).value))
            except:
                if ws.cell(row=i, column=col).value[:-1] != "" and ws.cell(row=i, column=col).value[:-1] != "-":
                    temp["marks"].append(
                        int(ws.cell(row=i, column=col).value[:-1]))
                else:
                    temp["marks"].append(0)
            temp['GPA'].append(ws.cell(row=i, column=82).value)
        temp = pd.DataFrame(temp)
        temp = temp.sort_values(
            by=["marks", "GPA"], ignore_index=True, ascending=False)
        i = 0
        while(temp["marks"].iloc[i] == temp["marks"].iloc[0] and i < temp['marks'].count()-1):
            subjectRankers["seat"].append(temp["seat"].iloc[i])
            subjectRankers["name"].append(temp["name"].iloc[i])
            subjectRankers["subject"].append(subjects[col])
            i += 1

    return subjectRankers


def topRankers(wb, source, class_name):
    ws = wb[source]
    ws2 = wb[class_name]
    top10 = findTop10Rankers(ws)
    ws2["E47"] = "Rankers"
    ws2['E47'].font = Font(size=14, bold=True)
    ws2['E47'].alignment = Alignment(horizontal='center')

    ws2["A48"], ws2["B48"], ws2["C48"], ws2["h48"], ws2["I48"] = "Topper", "Seat", "Name", "Total", "GPA"
    for i in range(0, 10):
        ws2["A"+str(49+i)].value = i+1
        ws2["B"+str(49+i)].value = top10["seat"][i]
        ws2["C"+str(49+i)].value = top10["name"][i]
        ws2["H"+str(49+i)].value = top10["total"][i]+"/"+top10["outof"][i]
        ws2["I"+str(49+i)].value = top10["GPA"][i]
    ws2.merge_cells('C48:G48')
    ws2.merge_cells('C49:G49')
    ws2.merge_cells('C50:G50')
    ws2.merge_cells('C51:G51')
    ws2.merge_cells('C52:G52')
    ws2.merge_cells('C53:G53')
    ws2.merge_cells('C54:G54')
    ws2.merge_cells('C55:G55')
    ws2.merge_cells('C56:G56')
    ws2.merge_cells('C57:G57')
    ws2.merge_cells('C58:G58')

    # subject rankers
    ws2["E80"] = "Rankers"
    ws2['E80'].font = Font(size=14, bold=True)
    ws2['E80'].alignment = Alignment(horizontal='center')
    subjectrankers = subjectRankers([7, 20, 34, 48, 61, 68, 74], {
                                    7: "MFCS", 20: "AJAVA", 34: "ADBMS", 48: "SPM", 61: 'DSL', 68: 'WTL', 74: 'MP'}, ws)
    ws2["A81"], ws2["B81"], ws2["C81"], ws2["h81"] = "Topper", "Seat", "Name", "Subject"
    for i in range(0, len(subjectrankers["seat"])):
        ws2["A"+str(82+i)].value = i+1
        ws2["B"+str(82+i)].value = int(subjectrankers["seat"][i])
        ws2["C"+str(82+i)].value = subjectrankers["name"][i]
        ws2["H"+str(82+i)].value = subjectrankers["subject"][i]
        ws2.merge_cells("C"+str(82+i)+":G"+str(82+i))

    return wb


def barChart(wb, class_name):
    ws2 = wb[class_name]
    chart1 = BarChart()
    chart1.type = "col"
    chart1.style = 10
    chart1.title = "Bar Chart"
    chart1.x_axis.title = 'Subjects'

    cats = Reference(ws2, min_col=2, min_row=41, max_col=12, max_row=41)

    series1 = Reference(ws2, min_col=1, min_row=42, max_col=12, max_row=42)
    chart1.series.append(Series(series1, title_from_data=True))

    series2 = Reference(ws2, min_col=1, min_row=43, max_col=12, max_row=43)
    chart1.series.append(Series(series2, title_from_data=True))

    # chart1.add_data(data)
    chart1.set_categories(cats)
    chart1.y_axis.scaling.min = 0
    chart1.y_axis.scaling.max = 100

    chart1.shape = 4
    ws2.add_chart(chart1, "C62")
    return wb


def genrateAnalysis_MCAsem1(path, source, class_name):
    wb = load_workbook(path)
    wb.create_sheet(class_name)
    wb = create_Boilerplate(wb, class_name)
    wb = overall_summary_of_the_semester(wb, source, class_name)
    wb = topRankers(wb, source, class_name)
    wb = barChart(wb, class_name)
    wb.save(path)
