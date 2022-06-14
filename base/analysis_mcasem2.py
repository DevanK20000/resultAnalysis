from numpy import double
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.chart import BarChart, Reference, Series
import pandas as pd


def create_Boilerplate(wb, class_name):
    ws = wb[class_name]
    ws.column_dimensions['A'].width = 40

    ws.merge_cells('A2:H2')
    ws['A2'] = 'Vivekanand Education Society Institute of Technology'

    ws.merge_cells('A3:H3')
    ws['A3'] = 'Master of Computer Applications'

    ws.merge_cells('A4:H4')
    ws['A4'] = 'Result Analysis –Ist year-Sem 2'

    ws.merge_cells('A7:H7')
    ws['A7'] = 'OVERALL SUMMARY OF THE SEMESTER'

    ws['A2'].font = Font(size=12, bold=True)
    ws['A3'].font = Font(size=12, bold=True)
    ws['A4'].font = Font(size=12, bold=True)
    ws['A7'].font = Font(size=14, bold=True)

    ws['A2'].alignment = Alignment(horizontal='center')
    ws['A3'].alignment = Alignment(horizontal='center')
    ws['A4'].alignment = Alignment(horizontal='center')
    ws['A7'].alignment = Alignment(horizontal='center')

    return wb


def getAverage(column, ws):
    average = 0
    for i in range(2, ws.max_row+1):
        try:
            average += int(ws.cell(row=i, column=column).value)
        except:
            if ws.cell(row=i, column=column).value != "--" and ws.cell(row=i, column=column).value != "A":
                average += int((ws.cell(row=i, column=column).value)[:-1])
    average = round(average/(ws.max_row-1))
    return average


def getAverageElective(column, elective_column, elective_name, ws):
    average = 0
    for i in range(2, ws.max_row+1):
        if (ws.cell(row=i, column=elective_column).value).find(elective_name) != -1:
            try:
                average += int(ws.cell(row=i, column=column).value)
            except:
                if ws.cell(row=i, column=column).value != "--":
                    average += int((ws.cell(row=i, column=column).value)[:-1])
    if int(checkElectiveCount(elective_name, elective_column, ws)) == 0:
        return average
    else:
        average = round(
            average/checkElectiveCount(elective_name, elective_column, ws))
        return average


def getPassed(column, ws):
    count = 0
    for i in range(2, ws.max_row+1):
        if str(ws.cell(row=i, column=column).value) != "--" and str(ws.cell(row=i, column=column).value) != "A" and str(ws.cell(row=i, column=column).value).find("F") == -1:
            count += 1
    return count


def getPassedElective(column, elective_column, elective_name, ws):
    count = 0
    for i in range(2, ws.max_row+1):
        if (ws.cell(row=i, column=elective_column).value).find(elective_name) != -1:
            if str(ws.cell(row=i, column=column).value) != "--" and str(ws.cell(row=i, column=column).value) != "A" and str(ws.cell(row=i, column=column).value).find("F") == -1:
                count += 1
    return count


def getTotalPFA(column, g, ws):
    total = 0
    for i in range(2, ws.max_row+1):
        if ws.cell(row=i, column=column).value == g:
            total += 1
    return total


def getGradeCount(column, grade, ws):
    count = 0
    for i in range(2, ws.max_row+1):
        if ws.cell(row=i, column=column).value == grade or (grade == "F" and ws.cell(row=i, column=column).value == "--"):
            count += 1
    return count


def getGradeCountElective(column, elective_column, elective_name, grade, ws):
    count = 0
    for i in range(2, ws.max_row+1):
        if (ws.cell(row=i, column=elective_column).value).find(elective_name) != -1:
            if ws.cell(row=i, column=column).value == grade or (grade == "F" and ws.cell(row=i, column=column).value == "--"):
                count += 1
    return count


def percentageOfStudentsWith60amdAbovePercentage(column, ws):
    count = 0
    for i in range(2, ws.max_row+1):
        if (ws.cell(row=i, column=column).value) != "--":
            if double(ws.cell(row=i, column=column).value)*7.4+12 >= 60:
                count += 1
    return (count/(ws.max_row-1))*100


def percentageOfStudentsBelow60(column, ws):
    count = 0
    for i in range(2, ws.max_row+1):
        if (ws.cell(row=i, column=column).value) != "--":
            if double(ws.cell(row=i, column=column).value)*7.4+12 < 60:
                count += 1
    return count/(ws.max_row-1)*100


def checkElectiveCount(name, column, ws):
    count = 0
    for i in range(2, ws.max_row+1):
        if (ws.cell(row=i, column=column).value).find(name) != -1:
            count += 1
    return count


def checkAppeared(column, ws):
    count = 0
    for i in range(2, ws.max_row+1):
        if str(ws.cell(row=i, column=column).value).find("A") == -1:
            count += 1
    return count


def checkAppearedElective(column, elective_name, elective_column, ws):
    count = 0
    for i in range(2, ws.max_row+1):
        if (ws.cell(row=i, column=elective_column).value).find(elective_name) != -1:
            if str(ws.cell(row=i, column=column).value).find("A") == -1:
                count += 1
    return count


def overall_summary_of_the_semester(wb, source, class_name):
    ws = wb[source]
    ws2 = wb[class_name]

    ws2.append(['Subjects', 'MFCS2', 'MFCS2-TW', 'AIML', 'AIML', 'IS', 'IP', 'IPL', 'IOT',
               'IOTL', 'NLP', 'NLPL', 'DMBA', 'DMBAL', 'SSDL', 'AWTL', 'UIL', 'NL', 'MP', 'Total'])
    ws2.append(['Marks', 100, 25, 100, 75, 100, 100, 75, 100,
               25, 100, 25, 100, 25, 50, 100, 75, 75, 50, 1050])
    ws2.append(['Total Students',
                ws.max_row-1, ws.max_row-1,
                ws.max_row-1, ws.max_row-1,
                ws.max_row-1,
                checkElectiveCount("IMAGE PROCESSING", 2, ws), checkElectiveCount(
                    "IMAGE PROCESSING", 2, ws),
                checkElectiveCount("INTERNET OF THINGS", 2, ws), checkElectiveCount(
                    "INTERNET OF THINGS", 2, ws),
                checkElectiveCount("NATURAL LANGUAGE PROCESSING", 3, ws), checkElectiveCount(
                    "NATURAL LANGUAGE PROCESSING", 3, ws),
                checkElectiveCount("DIGITAL MARKETING AND BUSINESS ANALYTICS", 3, ws), checkElectiveCount(
                    "DIGITAL MARKETING AND BUSINESS ANALYTICS", 3, ws),
                ws.max_row-1, ws.max_row-1, ws.max_row-1, ws.max_row-1, ws.max_row-1
                ])
    # ws2["T10"].value=ws2["B10"].value+ws2["C10"].value+ws2["D10"].value+ws2["F10"].value+ws2["G10"].value+ws2["H10"].value+ws2["I10"].value+ws2["J10"].value+ws2["K10"].value+ws2["L10"].value+ws2["M10"].value+ws2["N10"].value+ws2["O10"].value+ws2["P10"].value+ws2["Q10"].value+ws2["R10"].value+ws2["S10"].value

    ws2.append(['Total Apeared', checkAppeared(8, ws), checkAppeared(14, ws), checkAppeared(21, ws), checkAppeared(28, ws), checkAppeared(35, ws),
                checkAppearedElective(42, "IMAGE PROCESSING", 2, ws), checkAppearedElective(
                    49, "IMAGE PROCESSING", 2, ws),
                checkAppearedElective(42, "INTERNET OF THINGS", 2, ws), checkAppearedElective(
                    49, "INTERNET OF THINGS", 2, ws),
                checkAppearedElective(56, "NATURAL LANGUAGE PROCESSING", 3, ws), checkAppearedElective(
                    62, "NATURAL LANGUAGE PROCESSING", 3, ws),
                checkAppearedElective(56, "DIGITAL MARKETING AND BUSINESS ANALYTICS", 3, ws), checkAppearedElective(
                    62, "DIGITAL MARKETING AND BUSINESS ANALYTICS", 3, ws),
                checkAppeared(68, ws), checkAppeared(74, ws), checkAppeared(81, ws), checkAppeared(88, ws), checkAppeared(94, ws)])

    # ws2["T11"].value=ws2["B11"].value+ws2["C11"].value+ws2["D11"].value+ws2["F11"].value+ws2["G11"].value+ws2["H11"].value+ws2["I11"].value+ws2["J11"].value+ws2["K11"].value+ws2["L11"].value+ws2["M11"].value+ws2["N11"].value+ws2["O11"].value+ws2["P11"].value+ws2["Q11"].value+ws2["R11"].value+ws2["S11"].value

    ws2.append(['Average Marks', getAverage(9, ws), getAverage(15, ws), getAverage(22, ws), getAverage(29, ws), getAverage(36, ws),
                getAverageElective(43, 2, "IMAGE PROCESSING", ws), getAverageElective(
                    50, 2, "IMAGE PROCESSING", ws),
                getAverageElective(43, 2, "INTERNET OF THINGS", ws), getAverageElective(
                    50, 2, "INTERNET OF THINGS", ws),
                getAverageElective(57, 3, "NATURAL LANGUAGE PROCESSING", ws), getAverageElective(
                    63, 3, "NATURAL LANGUAGE PROCESSING", ws),
                getAverageElective(57, 3, "DIGITAL MARKETING AND BUSINESS ANALYTICS", ws), getAverageElective(
                    63, 3, "DIGITAL MARKETING AND BUSINESS ANALYTICS", ws),
                getAverage(68, ws),
                getAverage(75, ws),
                getAverage(82, ws),
                getAverage(89, ws),
                getAverage(94, ws)])
    ws2.append(["Total Passed", getPassed(9, ws), getPassed(15, ws), getPassed(22, ws), getPassed(29, ws), getPassed(36, ws),
                getPassedElective(43, 2, "IMAGE PROCESSING", ws), getPassedElective(
                    50, 2, "IMAGE PROCESSING", ws),
                getPassedElective(43, 2, "INTERNET OF THINGS", ws), getPassedElective(
                    50, 2, "INTERNET OF THINGS", ws),
                getPassedElective(57, 3, "NATURAL LANGUAGE PROCESSING", ws),
                getPassedElective(63, 3, "NATURAL LANGUAGE PROCESSING", ws),
                getPassedElective(
                    57, 3, "DIGITAL MARKETING AND BUSINESS ANALYTICS", ws),
                getPassedElective(
                    63, 3, "DIGITAL MARKETING AND BUSINESS ANALYTICS", ws),
                getPassed(68, ws),
                getPassed(75, ws),
                getPassed(82, ws),
                getPassed(89, ws),
                getPassed(94, ws)])
    ws2['T13'] = getTotalPFA(103, "P", ws)
    ws2.append(["Total Failed", (ws2["B11"].value-ws2["B13"].value), (ws2["C11"].value-ws2["C13"].value), (ws2["D11"].value-ws2["D13"].value), (ws2["E11"].value-ws2["E13"].value), (ws2["F11"].value-ws2["F13"].value), (ws2["G11"].value-ws2["G13"].value), (ws2["H11"].value-ws2["H13"].value), (ws2["I11"].value-ws2["I13"].value), (ws2["J11"].value -
               ws2["J13"].value), (ws2["K11"].value-ws2["K13"].value), (ws2["L11"].value-ws2["L13"].value), (ws2["M11"].value-ws2["M13"].value), (ws2["N11"].value-ws2["N13"].value), (ws2["O11"].value-ws2["O13"].value), (ws2["P11"].value-ws2["P13"].value), (ws2["Q11"].value-ws2["Q13"].value), (ws2["R11"].value-ws2["R13"].value), (ws2["S11"].value-ws2["S13"].value)])
    ws2.append(["Total Absent", (ws2["B10"].value-ws2["B11"].value), (ws2["C10"].value-ws2["C11"].value), (ws2["D10"].value-ws2["D11"].value), (ws2["E10"].value-ws2["E11"].value), (ws2["F10"].value-ws2["F11"].value), (ws2["G10"].value-ws2["G11"].value), (ws2["H10"].value-ws2["H11"].value), (ws2["I10"].value-ws2["I11"].value), (ws2["J10"].value -
               ws2["J11"].value), (ws2["K10"].value-ws2["K11"].value), (ws2["L10"].value-ws2["L11"].value), (ws2["M10"].value-ws2["M11"].value), (ws2["N10"].value-ws2["N11"].value), (ws2["O10"].value-ws2["O11"].value), (ws2["P10"].value-ws2["P11"].value), (ws2["Q10"].value-ws2["Q11"].value), (ws2["R10"].value-ws2["R11"].value), (ws2["S10"].value-ws2["S11"].value)])

    ws2.append(["Percentage Passed",
                (ws2['B13'].value/ws2["B10"].value *
                 100) if ws2["B10"].value != 0 else 0,
                (ws2['C13'].value/ws2["C10"].value *
                 100) if ws2["C10"].value != 0 else 0,
                (ws2['D13'].value/ws2["D10"].value *
                 100) if ws2["D10"].value != 0 else 0,
                (ws2['E13'].value/ws2["E10"].value *
                 100) if ws2["E10"].value != 0 else 0,
                (ws2['F13'].value/ws2["F10"].value *
                 100) if ws2["F10"].value != 0 else 0,
                (ws2['G13'].value/ws2["G10"].value *
                 100) if ws2["G10"].value != 0 else 0,
                (ws2['H13'].value/ws2["H10"].value *
                 100) if ws2["H10"].value != 0 else 0,
                (ws2['I13'].value/ws2["I10"].value *
                 100) if ws2["I10"].value != 0 else 0,
                (ws2['J13'].value/ws2["J10"].value *
                 100) if ws2["J10"].value != 0 else 0,
                (ws2['K13'].value/ws2["K10"].value *
                 100) if ws2["K10"].value != 0 else 0,
                (ws2['L13'].value/ws2["L10"].value *
                 100) if ws2["L10"].value != 0 else 0,
                (ws2['M13'].value/ws2["M10"].value *
                 100) if ws2["M10"].value != 0 else 0,
                (ws2['N13'].value/ws2["N10"].value *
                 100) if ws2["N10"].value != 0 else 0,
                (ws2['O13'].value/ws2["O10"].value *
                 100) if ws2["O10"].value != 0 else 0,
                (ws2['P13'].value/ws2["P10"].value *
                 100) if ws2["P10"].value != 0 else 0,
                (ws2['Q13'].value/ws2["Q10"].value *
                 100) if ws2["Q10"].value != 0 else 0,
                (ws2['R13'].value/ws2["R10"].value *
                 100) if ws2["R10"].value != 0 else 0,
                (ws2['S13'].value/ws2["S10"].value*100) if ws2["S10"].value != 0 else 0])
    # ws2['T16'] = ws2['T14'].value/(ws.max_row-1)*100

    ws2.append(["Percentage Failed",
                (ws2['B14'].value/ws2["B10"].value *
                 100)if ws2["B10"].value != 0 else 0,
                (ws2['C14'].value/ws2["C10"].value *
                 100)if ws2["C10"].value != 0 else 0,
                (ws2['D14'].value/ws2["D10"].value *
                 100)if ws2["D10"].value != 0 else 0,
                (ws2['E14'].value/ws2["E10"].value *
                 100)if ws2["E10"].value != 0 else 0,
                (ws2['F14'].value/ws2["F10"].value *
                 100)if ws2["F10"].value != 0 else 0,
                (ws2['G14'].value/ws2["G10"].value *
                 100)if ws2["G10"].value != 0 else 0,
                (ws2['H14'].value/ws2["H10"].value *
                 100)if ws2["H10"].value != 0 else 0,
                (ws2['I14'].value/ws2["I10"].value *
                 100)if ws2["I10"].value != 0 else 0,
                (ws2['J14'].value/ws2["J10"].value *
                 100)if ws2["J10"].value != 0 else 0,
                (ws2['K14'].value/ws2["K10"].value *
                 100)if ws2["K10"].value != 0 else 0,
                (ws2['L14'].value/ws2["L10"].value *
                 100)if ws2["L10"].value != 0 else 0,
                (ws2['M14'].value/ws2["M10"].value *
                 100)if ws2["M10"].value != 0 else 0,
                (ws2['N14'].value/ws2["N10"].value *
                 100)if ws2["N10"].value != 0 else 0,
                (ws2['O14'].value/ws2["O10"].value *
                 100)if ws2["O10"].value != 0 else 0,
                (ws2['P14'].value/ws2["P10"].value *
                 100)if ws2["P10"].value != 0 else 0,
                (ws2['Q14'].value/ws2["Q10"].value *
                 100)if ws2["Q10"].value != 0 else 0,
                (ws2['R14'].value/ws2["R10"].value *
                 100)if ws2["R10"].value != 0 else 0,
                (ws2['S14'].value/ws2["S10"].value*100)if ws2["S10"].value != 0 else 0])
    # ws2['T17'] = ws2['T15'].value/(ws.max_row-1)*100
    ws2.append(["Number of students with >= 80% ie =Grade O", getGradeCount(11, 'O', ws), getGradeCount(17, 'O', ws), getGradeCount(24, 'O', ws), getGradeCount(31, 'O', ws), getGradeCount(38, 'O', ws),
                getGradeCountElective(45, 2, "IMAGE PROCESSING", "O", ws), getGradeCountElective(
                    52, 2, "IMAGE PROCESSING", "O", ws),
                getGradeCountElective(45, 2, "INTERNET OF THINGS", "O", ws), getGradeCountElective(
                    52, 2, "INTERNET OF THINGS", "O", ws),
                getGradeCountElective(
                    59, 3, "NATURAL LANGUAGE PROCESSING", "O", ws),
                getGradeCountElective(
                    65, 3, "NATURAL LANGUAGE PROCESSING", "O", ws),
                getGradeCountElective(
                    59, 3, "DIGITAL MARKETING AND BUSINESS ANALYTICS", "O", ws),
                getGradeCountElective(
                    65, 3, "DIGITAL MARKETING AND BUSINESS ANALYTICS", "O", ws),
                getGradeCount(70, "O", ws),
                getGradeCount(77, "O", ws),
                getGradeCount(84, "O", ws),
                getGradeCount(91, "O", ws),
                getGradeCount(96, "O", ws)])
    ws2.append(["Number of students with >= 75% ie =Grade A", getGradeCount(11, 'A', ws), getGradeCount(17, 'A', ws), getGradeCount(24, 'A', ws), getGradeCount(31, 'A', ws), getGradeCount(38, 'A', ws),
                getGradeCountElective(45, 2, "IMAGE PROCESSING", 'A', ws), getGradeCountElective(
                    52, 2, "IMAGE PROCESSING", 'A', ws),
                getGradeCountElective(45, 2, "INTERNET OF THINGS", 'A', ws), getGradeCountElective(
                    52, 2, "INTERNET OF THINGS", 'A', ws),
                getGradeCountElective(
                    59, 3, "NATURAL LANGUAGE PROCESSING", 'A', ws),
                getGradeCountElective(
                    65, 3, "NATURAL LANGUAGE PROCESSING", 'A', ws),
                getGradeCountElective(
                    59, 3, "DIGITAL MARKETING AND BUSINESS ANALYTICS", 'A', ws),
                getGradeCountElective(
                    65, 3, "DIGITAL MARKETING AND BUSINESS ANALYTICS", 'A', ws),
                getGradeCount(70, 'A', ws),
                getGradeCount(77, 'A', ws),
                getGradeCount(84, 'A', ws),
                getGradeCount(91, 'A', ws),
                getGradeCount(96, 'A', ws)])
    ws2.append(["Number of students with >= 70% ie =Grade B", getGradeCount(11, "B", ws), getGradeCount(17, "B", ws), getGradeCount(24, "B", ws), getGradeCount(31, "B", ws), getGradeCount(38, "B", ws),
                getGradeCountElective(45, 2, "IMAGE PROCESSING", "B", ws), getGradeCountElective(
                    52, 2, "IMAGE PROCESSING", "B", ws),
                getGradeCountElective(45, 2, "INTERNET OF THINGS", "B", ws), getGradeCountElective(
                    52, 2, "INTERNET OF THINGS", "B", ws),
                getGradeCountElective(
                    59, 3, "NATURAL LANGUAGE PROCESSING", "B", ws),
                getGradeCountElective(
                    65, 3, "NATURAL LANGUAGE PROCESSING", "B", ws),
                getGradeCountElective(
                    59, 3, "DIGITAL MARKETING AND BUSINESS ANALYTICS", "B", ws),
                getGradeCountElective(
                    65, 3, "DIGITAL MARKETING AND BUSINESS ANALYTICS", "B", ws),
                getGradeCount(70, "B", ws),
                getGradeCount(77, "B", ws),
                getGradeCount(84, "B", ws),
                getGradeCount(91, "B", ws),
                getGradeCount(96, "B", ws)])
    ws2.append(["Number of students with >= 60% ie =Grade C", getGradeCount(11, "C", ws), getGradeCount(17, "C", ws), getGradeCount(24, "C", ws), getGradeCount(31, "C", ws), getGradeCount(38, "C", ws),
                getGradeCountElective(45, 2, "IMAGE PROCESSING", "C", ws), getGradeCountElective(
                    52, 2, "IMAGE PROCESSING", "C", ws),
                getGradeCountElective(45, 2, "INTERNET OF THINGS", "C", ws), getGradeCountElective(
                    52, 2, "INTERNET OF THINGS", "C", ws),
                getGradeCountElective(
                    59, 3, "NATURAL LANGUAGE PROCESSING", "C", ws),
                getGradeCountElective(
                    65, 3, "NATURAL LANGUAGE PROCESSING", "C", ws),
                getGradeCountElective(
                    59, 3, "DIGITAL MARKETING AND BUSINESS ANALYTICS", "C", ws),
                getGradeCountElective(
                    65, 3, "DIGITAL MARKETING AND BUSINESS ANALYTICS", "C", ws),
                getGradeCount(70, "C", ws),
                getGradeCount(77, "C", ws),
                getGradeCount(84, "C", ws),
                getGradeCount(91, "C", ws),
                getGradeCount(96, "C", ws)])
    ws2.append(["Number of students with >= 55% ie =Grade D", getGradeCount(11, "D", ws), getGradeCount(17, "D", ws), getGradeCount(24, "D", ws), getGradeCount(31, "D", ws), getGradeCount(38, "D", ws),
                getGradeCountElective(45, 2, "IMAGE PROCESSING", "D", ws), getGradeCountElective(
                    52, 2, "IMAGE PROCESSING", "D", ws),
                getGradeCountElective(45, 2, "INTERNET OF THINGS", "D", ws), getGradeCountElective(
                    52, 2, "INTERNET OF THINGS", "D", ws),
                getGradeCountElective(
                    59, 3, "NATURAL LANGUAGE PROCESSING", "D", ws),
                getGradeCountElective(
                    65, 3, "NATURAL LANGUAGE PROCESSING", "D", ws),
                getGradeCountElective(
                    59, 3, "DIGITAL MARKETING AND BUSINESS ANALYTICS", "D", ws),
                getGradeCountElective(
                    65, 3, "DIGITAL MARKETING AND BUSINESS ANALYTICS", "D", ws),
                getGradeCount(70, "D", ws),
                getGradeCount(77, "D", ws),
                getGradeCount(84, "D", ws),
                getGradeCount(91, "D", ws),
                getGradeCount(96, "D", ws)])
    ws2.append(["Number of students with >= 50% ie =Grade E", getGradeCount(11, "E", ws), getGradeCount(17, "E", ws), getGradeCount(24, "E", ws), getGradeCount(31, "E", ws), getGradeCount(38, "E", ws),
                getGradeCountElective(45, 2, "IMAGE PROCESSING", "E", ws), getGradeCountElective(
                    52, 2, "IMAGE PROCESSING", "E", ws),
                getGradeCountElective(45, 2, "INTERNET OF THINGS", "E", ws), getGradeCountElective(
                    52, 2, "INTERNET OF THINGS", "E", ws),
                getGradeCountElective(
                    59, 3, "NATURAL LANGUAGE PROCESSING", "E", ws),
                getGradeCountElective(
                    65, 3, "NATURAL LANGUAGE PROCESSING", "E", ws),
                getGradeCountElective(
                    59, 3, "DIGITAL MARKETING AND BUSINESS ANALYTICS", "E", ws),
                getGradeCountElective(
                    65, 3, "DIGITAL MARKETING AND BUSINESS ANALYTICS", "E", ws),
                getGradeCount(70, "E", ws),
                getGradeCount(77, "E", ws),
                getGradeCount(84, "E", ws),
                getGradeCount(91, "E", ws),
                getGradeCount(96, "E", ws)])
    ws2.append(["Number of students with >= 45% ie =Grade P", getGradeCount(11, "P", ws), getGradeCount(17, "P", ws), getGradeCount(24, "P", ws), getGradeCount(31, "P", ws), getGradeCount(38, "P", ws),
                getGradeCountElective(45, 2, "IMAGE PROCESSING", "P", ws), getGradeCountElective(
                    52, 2, "IMAGE PROCESSING", "P", ws),
                getGradeCountElective(45, 2, "INTERNET OF THINGS", "P", ws), getGradeCountElective(
                    52, 2, "INTERNET OF THINGS", "P", ws),
                getGradeCountElective(
                    59, 3, "NATURAL LANGUAGE PROCESSING", "P", ws),
                getGradeCountElective(
                    65, 3, "NATURAL LANGUAGE PROCESSING", "P", ws),
                getGradeCountElective(
                    59, 3, "DIGITAL MARKETING AND BUSINESS ANALYTICS", "P", ws),
                getGradeCountElective(
                    65, 3, "DIGITAL MARKETING AND BUSINESS ANALYTICS", "P", ws),
                getGradeCount(70, "P", ws),
                getGradeCount(77, "P", ws),
                getGradeCount(84, "P", ws),
                getGradeCount(91, "P", ws),
                getGradeCount(96, "P", ws)])
    ws2.append(["Number of students with <= 45% ie =Grade F", getGradeCount(11, "F", ws), getGradeCount(17, "F", ws), getGradeCount(24, "F", ws), getGradeCount(31, "F", ws), getGradeCount(38, "F", ws),
                getGradeCountElective(45, 2, "IMAGE PROCESSING", "F", ws), getGradeCountElective(
                    52, 2, "IMAGE PROCESSING", "F", ws),
                getGradeCountElective(45, 2, "INTERNET OF THINGS", "F", ws), getGradeCountElective(
                    52, 2, "INTERNET OF THINGS", "F", ws),
                getGradeCountElective(
                    59, 3, "NATURAL LANGUAGE PROCESSING", "F", ws),
                getGradeCountElective(
                    65, 3, "NATURAL LANGUAGE PROCESSING", "F", ws),
                getGradeCountElective(
                    59, 3, "DIGITAL MARKETING AND BUSINESS ANALYTICS", "F", ws),
                getGradeCountElective(
                    65, 3, "DIGITAL MARKETING AND BUSINESS ANALYTICS", "F", ws),
                getGradeCount(70, "F", ws),
                getGradeCount(77, "F", ws),
                getGradeCount(84, "F", ws),
                getGradeCount(91, "F", ws),
                getGradeCount(96, "F", ws)])
    ws2.append(["Total Number of students with 60% and above",
                ws2['B18'].value+ws2['B19'].value +
                ws2['B20'].value+ws2['B21'].value,
                ws2['C18'].value+ws2['C19'].value +
                ws2['C20'].value+ws2['C21'].value,
                ws2['D18'].value+ws2['D19'].value +
                ws2['D20'].value+ws2['D21'].value,
                ws2['E18'].value+ws2['E19'].value +
                ws2['E20'].value+ws2['E21'].value,
                ws2['F18'].value+ws2['F19'].value +
                ws2['F20'].value+ws2['F21'].value,
                ws2['G18'].value+ws2['G19'].value +
                ws2['G20'].value+ws2['G21'].value,
                ws2['H18'].value+ws2['H19'].value +
                ws2['H20'].value+ws2['H21'].value,
                ws2['I18'].value+ws2['I19'].value +
                ws2['I20'].value+ws2['I21'].value,
                ws2['J18'].value+ws2['J19'].value +
                ws2['J20'].value+ws2['J21'].value,
                ws2['K18'].value+ws2['K19'].value +
                ws2['K20'].value+ws2['K21'].value,
                ws2['L18'].value+ws2['L19'].value +
                ws2['L20'].value+ws2['L21'].value,
                ws2['M18'].value+ws2['M19'].value +
                ws2['M20'].value+ws2['M21'].value,
                ws2['N18'].value+ws2['N19'].value +
                ws2['N20'].value+ws2['N21'].value,
                ws2['O18'].value+ws2['O19'].value +
                ws2['O20'].value+ws2['O21'].value,
                ws2['P18'].value+ws2['P19'].value +
                ws2['P20'].value+ws2['P21'].value,
                ws2['Q18'].value+ws2['Q19'].value +
                ws2['Q20'].value+ws2['Q21'].value,
                ws2['R18'].value+ws2['R19'].value +
                ws2['R20'].value+ws2['R21'].value,
                ws2['S18'].value+ws2['S19'].value+ws2['S20'].value+ws2['S21'].value])
    ws2.append(["% of students with 60% and above",
                (ws2["B26"].value/ws2["B10"].value *
                 100)if ws2["B10"].value != 0 else 0,
                (ws2["C26"].value/ws2["C10"].value *
                 100)if ws2["C10"].value != 0 else 0,
                (ws2["D26"].value/ws2["D10"].value *
                 100)if ws2["D10"].value != 0 else 0,
                (ws2["E26"].value/ws2["E10"].value *
                 100)if ws2["E10"].value != 0 else 0,
                (ws2["F26"].value/ws2["F10"].value *
                 100)if ws2["F10"].value != 0 else 0,
                (ws2["G26"].value/ws2["G10"].value *
                 100)if ws2["G10"].value != 0 else 0,
                (ws2["H26"].value/ws2["H10"].value *
                 100)if ws2["H10"].value != 0 else 0,
                (ws2["I26"].value/ws2["I10"].value *
                 100)if ws2["I10"].value != 0 else 0,
                (ws2["J26"].value/ws2["J10"].value *
                 100)if ws2["J10"].value != 0 else 0,
                (ws2["K26"].value/ws2["K10"].value *
                 100)if ws2["K10"].value != 0 else 0,
                (ws2["L26"].value/ws2["L10"].value *
                 100)if ws2["L10"].value != 0 else 0,
                (ws2["M26"].value/ws2["M10"].value *
                 100)if ws2["M10"].value != 0 else 0,
                (ws2["N26"].value/ws2["N10"].value *
                 100)if ws2["N10"].value != 0 else 0,
                (ws2["O26"].value/ws2["O10"].value *
                 100)if ws2["O10"].value != 0 else 0,
                (ws2["P26"].value/ws2["P10"].value *
                 100)if ws2["P10"].value != 0 else 0,
                (ws2["Q26"].value/ws2["Q10"].value *
                 100)if ws2["Q10"].value != 0 else 0,
                (ws2["R26"].value/ws2["R10"].value *
                 100)if ws2["R10"].value != 0 else 0,
                (ws2["S26"].value/ws2["S10"].value*100)if ws2["S10"].value != 0 else 0])

    ws2.append(["Total Number of students below 60%",
                ws2["B10"].value-ws2['B26'].value,
                ws2["C10"].value-ws2['C26'].value,
                ws2["D10"].value-ws2['D26'].value,
                ws2["E10"].value-ws2['E26'].value,
                ws2["F10"].value-ws2['F26'].value,
                ws2["G10"].value-ws2['G26'].value,
                ws2["H10"].value-ws2['H26'].value,
                ws2["I10"].value-ws2['I26'].value,
                ws2["J10"].value-ws2['J26'].value,
                ws2["K10"].value-ws2['K26'].value,
                ws2["L10"].value-ws2['L26'].value,
                ws2["M10"].value-ws2['M26'].value,
                ws2["N10"].value-ws2['N26'].value,
                ws2["O10"].value-ws2['O26'].value,
                ws2["P10"].value-ws2['P26'].value,
                ws2["Q10"].value-ws2['Q26'].value,
                ws2["R10"].value-ws2['R26'].value,
                ws2["S10"].value-ws2['S26'].value])
    ws2.append(["% of students below 60%",
                (ws2["B28"].value/ws2['B26'].value *
                 100)if ws2["B26"].value != 0 else 0,
                (ws2["C28"].value/ws2['C26'].value *
                 100)if ws2["C26"].value != 0 else 0,
                (ws2["D28"].value/ws2['D26'].value *
                 100)if ws2["D26"].value != 0 else 0,
                (ws2["E28"].value/ws2['E26'].value *
                 100)if ws2["E26"].value != 0 else 0,
                (ws2["F28"].value/ws2['F26'].value *
                 100)if ws2["F26"].value != 0 else 0,
                (ws2["G28"].value/ws2['G26'].value *
                 100)if ws2["G26"].value != 0 else 0,
                (ws2["H28"].value/ws2['H26'].value *
                 100)if ws2["H26"].value != 0 else 0,
                (ws2["I28"].value/ws2['I26'].value *
                 100)if ws2["I26"].value != 0 else 0,
                (ws2["J28"].value/ws2['J26'].value *
                 100)if ws2["J26"].value != 0 else 0,
                (ws2["K28"].value/ws2['K26'].value *
                 100)if ws2["K26"].value != 0 else 0,
                (ws2["L28"].value/ws2['L26'].value *
                 100)if ws2["L26"].value != 0 else 0,
                (ws2["M28"].value/ws2['M26'].value *
                 100)if ws2["M26"].value != 0 else 0,
                (ws2["N28"].value/ws2['N26'].value *
                 100)if ws2["N26"].value != 0 else 0,
                (ws2["O28"].value/ws2['O26'].value *
                 100)if ws2["O26"].value != 0 else 0,
                (ws2["P28"].value/ws2['P26'].value *
                 100)if ws2["P26"].value != 0 else 0,
                (ws2["Q28"].value/ws2['Q26'].value *
                 100)if ws2["Q26"].value != 0 else 0,
                (ws2["R28"].value/ws2['R26'].value *
                 100)if ws2["R26"].value != 0 else 0,
                (ws2["S28"].value/ws2['S26'].value*100)if ws2["S26"].value != 0 else 0])

    ws2.append(["Grand Total",
                ws2['B26'].value+ws2['B28'].value,
                ws2['C26'].value+ws2['C28'].value,
                ws2['D26'].value+ws2['D28'].value,
                ws2['E26'].value+ws2['E28'].value,
                ws2['F26'].value+ws2['F28'].value,
                ws2['G26'].value+ws2['G28'].value,
                ws2['H26'].value+ws2['H28'].value,
                ws2['I26'].value+ws2['I28'].value,
                ws2['J26'].value+ws2['J28'].value,
                ws2['K26'].value+ws2['K28'].value,
                ws2['L26'].value+ws2['L28'].value,
                ws2['M26'].value+ws2['M28'].value,
                ws2['N26'].value+ws2['N28'].value,
                ws2['O26'].value+ws2['O28'].value,
                ws2['P26'].value+ws2['P28'].value,
                ws2['Q26'].value+ws2['Q28'].value,
                ws2['R26'].value+ws2['R28'].value,
                ws2['S26'].value+ws2['S28'].value])

    ws2.append([""])
    ws2.append(["Total Student appeared", getTotalPFA(
        103, "P", ws)+getTotalPFA(103, "F", ws)])
    ws2.append(["passsed", getTotalPFA(103, "P", ws)])
    ws2.append(["failed", getTotalPFA(103, "F", ws)])
    ws2.append(["Absent", getTotalPFA(103, "A", ws)])
    ws2.append(["Total passing percentage",
               ws2['B33'].value/(ws.max_row-1)*100])
    ws2.append(["Total failing percentage",
               ws2['B34'].value/(ws.max_row-1)*100])
    ws2.append(["% of students with 60% and above",
               percentageOfStudentsWith60amdAbovePercentage(102, ws)])
    ws2.append(["% of students with below 60%", percentageOfStudentsBelow60(
        102, ws), "", "Total Absent percentage", "", ws2['B35'].value/(ws.max_row-1)*100])
    ws2.merge_cells('D39:E39')
    ws2.append([])
    ws2.append(['Subjects', 'MFCS2', 'MFCS2-TW', 'AIML', 'AIML', 'IS', 'IP', 'IPL', 'IOT',
               'IOTL', 'NLP', 'NLPL', 'DMBA', 'DMBAL', 'SSDL', 'AWTL', 'UIL', 'NL', 'MP', 'Total'])
    ws2.append(["% of students with 60% and above", ws2['B27'].value,
                ws2['C27'].value,
                ws2['D27'].value,
                ws2['E27'].value,
                ws2['F27'].value,
                ws2['G27'].value,
                ws2['H27'].value,
                ws2['I27'].value,
                ws2['J27'].value,
                ws2['K27'].value,
                ws2['L27'].value,
                ws2['M27'].value,
                ws2['N27'].value,
                ws2['O27'].value,
                ws2['P27'].value,
                ws2['Q27'].value,
                ws2['R27'].value,
                ws2['S27'].value])
    ws2.append(["% of students below 60%", ws2['B28'].value,
                ws2['C29'].value,
                ws2['D29'].value,
                ws2['E29'].value,
                ws2['F29'].value,
                ws2['G29'].value,
                ws2['H29'].value,
                ws2['I29'].value,
                ws2['J29'].value,
                ws2['K29'].value,
                ws2['L29'].value,
                ws2['M29'].value,
                ws2['N29'].value,
                ws2['O29'].value,
                ws2['P29'].value,
                ws2['Q29'].value,
                ws2['R29'].value,
                ws2['S29'].value])
    return wb


def findTop10Rankers(ws):
    top10 = {
        "seat": [],
        "name": [],
        "total": [],
        "outof": [],
        "GPA": [],
    }
    for i in range(2, ws.max_row):
        top10["seat"].append(ws.cell(row=i, column=4).value)
        top10["name"].append(ws.cell(row=i, column=5).value)
        top10["total"].append(ws.cell(row=i, column=99).value.split("/")[0])
        top10["outof"].append(ws.cell(row=i, column=99).value.split("/")[1])
        top10["GPA"].append(ws.cell(row=i, column=102).value)

    top10 = pd.DataFrame(top10)
    top10 = top10.sort_values(
        by=["total", "GPA"], ignore_index=True, ascending=False)
    return top10


def subjectRankers(column, subjects, ws):
    subjectRankers = {
        "seat": [],
        "name": [],
        "subject": [],
    }

    for col in column:
        temp = {
            "seat": [],
            "name": [],
            "marks": [],
            "GPA": [],
        }
        for i in range(2, ws.max_row):
            temp["seat"].append(ws.cell(row=i, column=4).value)
            temp["name"].append(ws.cell(row=i, column=5).value)
            try:
                temp["marks"].append(int(ws.cell(row=i, column=col).value))
            except:
                if ws.cell(row=i, column=col).value[:-1] != "" and ws.cell(row=i, column=col).value[:-1] != "-":
                    temp["marks"].append(
                        int(ws.cell(row=i, column=col).value[:-1]))
                else:
                    temp["marks"].append(0)
            temp['GPA'].append(ws.cell(row=i, column=82).value)
        temp = pd.DataFrame(temp)
        temp = temp.sort_values(
            by=["marks", "GPA"], ignore_index=True, ascending=False)
        i = 0
        while(temp["marks"].iloc[i] == temp["marks"].iloc[0] and i < temp['marks'].count()-1):
            subjectRankers["seat"].append(temp["seat"].iloc[i])
            subjectRankers["name"].append(temp["name"].iloc[i])
            subjectRankers["subject"].append(subjects[col])
            i += 1

    return subjectRankers


def subjectRankersElective(subjectRankers, column, subjects, elective_name, elective_column, ws):

    temp = {
        "elective1": [],
        "elective2": [],
        "seat": [],
        "name": [],
        "marks": [],
        "GPA": [],
    }
    for i in range(2, ws.max_row):

        temp["elective1"].append(ws.cell(row=i, column=2).value)
        temp["elective2"].append(ws.cell(row=i, column=3).value)
        temp["seat"].append(ws.cell(row=i, column=4).value)
        temp["name"].append(ws.cell(row=i, column=5).value)
        try:
            temp["marks"].append(int(ws.cell(row=i, column=column).value))
        except:
            if ws.cell(row=i, column=column).value[:-1] != "" and ws.cell(row=i, column=column).value[:-1] != "-":
                temp["marks"].append(
                    int(ws.cell(row=i, column=column).value[:-1]))
            else:
                temp["marks"].append(0)
        temp['GPA'].append(ws.cell(row=i, column=83).value)
    temp = pd.DataFrame(temp)

    for name in elective_name:
        temp2 = temp[temp["elective" +
                          str(elective_column-1)].str.contains(name)]
        temp2 = temp2.sort_values(
            by=["marks", "GPA"], ignore_index=True, ascending=False)
        i = 0
        try:
            while(temp2["marks"].iloc[i] == temp2["marks"].iloc[0] and i < temp2['marks'].count()-1):
                subjectRankers["seat"].append(temp2["seat"].iloc[i])
                subjectRankers["name"].append(temp2["name"].iloc[i])
                subjectRankers["subject"].append(subjects[name])
                i += 1
        except:
            pass
    return subjectRankers


def topRankers(wb, source, class_name):
    ws = wb[source]
    ws2 = wb[class_name]
    top10 = findTop10Rankers(ws)
    ws2["E47"] = "Rankers"
    ws2['E47'].font = Font(size=14, bold=True)
    ws2['E47'].alignment = Alignment(horizontal='center')

    ws2["A48"], ws2["B48"], ws2["C48"], ws2["h48"], ws2["I48"] = "Topper", "Seat", "Name", "Total", "GPA"
    for i in range(0, 10):
        ws2["A"+str(49+i)].value = i+1
        ws2["B"+str(49+i)].value = top10["seat"][i]
        ws2["C"+str(49+i)].value = top10["name"][i]
        ws2["H"+str(49+i)].value = top10["total"][i]+"/"+top10["outof"][i]
        ws2["I"+str(49+i)].value = top10["GPA"][i]
    ws2.merge_cells('C48:G48')
    ws2.merge_cells('C49:G49')
    ws2.merge_cells('C50:G50')
    ws2.merge_cells('C51:G51')
    ws2.merge_cells('C52:G52')
    ws2.merge_cells('C53:G53')
    ws2.merge_cells('C54:G54')
    ws2.merge_cells('C55:G55')
    ws2.merge_cells('C56:G56')
    ws2.merge_cells('C57:G57')
    ws2.merge_cells('C58:G58')

    # subject rankers
    ws2["E80"] = "Rankers"
    ws2['E80'].font = Font(size=14, bold=True)
    ws2['E80'].alignment = Alignment(horizontal='center')
    subjectrankers = subjectRankers(
        [9, 22, 36], {9: "MFCS2", 22: "AIML", 36: "IS"}, ws)
    subjectrankers = subjectRankersElective(subjectrankers, 36, {"IMAGE PROCESSING": "IP", "INTERNET OF THINGS": "IOT"}, [
                                            "IMAGE PROCESSING", "INTERNET OF THINGS"], 2, ws)
    subjectrankers = subjectRankersElective(subjectrankers, 43, {"DIGITAL MARKETING AND BUSINESS ANALYTICS": "DMBA", "NATURAL LANGUAGE PROCESSING": "NLP"}, [
                                            "DIGITAL MARKETING AND BUSINESS ANALYTICS", "NATURAL LANGUAGE PROCESSING"], 3, ws)
    ws2["A81"], ws2["B81"], ws2["C81"], ws2["h81"] = "Topper", "Seat", "Name", "Subject"
    for i in range(0, len(subjectrankers["seat"])):
        ws2["A"+str(82+i)].value = i+1
        ws2["B"+str(82+i)].value = int(subjectrankers["seat"][i])
        ws2["C"+str(82+i)].value = subjectrankers["name"][i]
        ws2["H"+str(82+i)].value = subjectrankers["subject"][i]
        ws2.merge_cells("C"+str(82+i)+":G"+str(82+i))

    return wb


def barChart(wb, class_name):
    ws2 = wb[class_name]
    chart1 = BarChart()
    chart1.type = "col"
    chart1.style = 10
    chart1.title = "Bar Chart"
    chart1.x_axis.title = 'Subjects'

    # data = Reference(ws2, min_col=1, min_row=42, max_col=12, max_row=43)
    cats = Reference(ws2, min_col=2, min_row=41, max_col=19, max_row=41)

    series1 = Reference(ws2, min_col=1, min_row=42, max_col=19, max_row=42)
    chart1.series.append(Series(series1, title_from_data=True))

    series2 = Reference(ws2, min_col=1, min_row=43, max_col=19, max_row=43)
    chart1.series.append(Series(series2, title_from_data=True))

    # chart1.add_data(data)
    chart1.set_categories(cats)

    chart1.y_axis.scaling.min = 0
    chart1.y_axis.scaling.max = 100

    chart1.shape = 4
    ws2.add_chart(chart1, "C62")
    return wb


def genrateAnalysis_MCAsem2(path, source, class_name):
    wb = load_workbook(path)
    wb.create_sheet(class_name)
    wb = create_Boilerplate(wb, class_name)
    wb = overall_summary_of_the_semester(wb, source, class_name)
    wb = topRankers(wb, source, class_name)
    wb = barChart(wb, class_name)
    wb.save(path)
